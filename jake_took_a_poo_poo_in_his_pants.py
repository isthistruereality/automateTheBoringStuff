from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

"""
# Uncomment this out when you want to dynamically populate the file to analyze
print('Please enter the file directory to the Excel file to analyze:')
file_dir = input()
"""

wb = load_workbook(filename = "Tufts Employee Census Data.xlsx")
sn = wb.sheetnames


def find_sheetname():
    sn = wb.sheetnames
    for i in sn:
        ws = wb[i]
        for row in ws.iter_rows(max_row=2,values_only=True):
            print(row)
            print("Is this the sheet you want?: (y/n)")
            a = input()
            if a.lower() == 'y':
                return wb[i]

#max_row=ws.max_row
#max_col=ws.max_column


def find_zip():
    position = 1
    
    for row in ws.values:
        for value in row:
        
            if 'zip' in str(value).lower():
                return get_column_letter(position)
            position += 1


def main():
    sheet_to_analyze = find_sheetname()
    print(sheet_to_analyze)
    #zip_col = find_zip()
    #print(zip_col)


main()
